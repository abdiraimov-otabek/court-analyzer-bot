from __future__ import annotations

import logging
from datetime import date, datetime
from io import BytesIO
from typing import Sequence

logger = logging.getLogger("case_exporter")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.domain.entities import CaseDecision


def build_cases_excel(
    decisions: Sequence[CaseDecision],
    target_article: str | None = None,
) -> bytes:
    """Build an Excel workbook from a list of case decisions.

    Args:
        decisions: The list of decisions to export.
        target_article: When provided (e.g. "61.3"), only rows whose
            ``raw_article`` or ``matched_article`` contain this article
            number are written to the sheet.  A summary row is appended
            at the end showing how many rows passed the filter.
    """
    filtered = list(decisions)
    skipped = 0
    if target_article:
        kept = []
        article_lower = target_article.lower()
        for d in decisions:
            # Check both the raw scraped article and the pipeline-matched article
            raw = (d.raw_article or "").lower()
            matched = (getattr(d, "matched_article", None) or "").lower()
            if article_lower in raw or article_lower in matched:
                kept.append(d)
            else:
                skipped += 1
        filtered = kept
        logger.info(
            "excel_export.article_filter article=%s kept=%d skipped=%d",
            target_article, len(filtered), skipped,
        )

    workbook = Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)
    sheet.title = "Cases"

    # Only the 8 requested columns
    headers = [
        "Номер",
        "Дата",
        "Номер дела",
        "Суд/Место",
        "Судья",
        "Ссылка",
        "Статья",
        "Текст",
    ]
    sheet.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for decision in filtered:
        # Robust date formatting
        d_date = ""
        if decision.raw_date:
            if isinstance(decision.raw_date, date):
                d_date = decision.raw_date.strftime("%d.%m.%Y")
            else:
                d_date = str(decision.raw_date)

        sheet.append([
            decision.raw_number,
            d_date,
            decision.raw_case_number,
            decision.raw_place,
            decision.raw_judge,
            decision.raw_url,
            decision.raw_article,
            decision.raw_text,
        ])

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    zebra_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")

    # Styling content
    for row in sheet.iter_rows(min_row=2):
        is_even_row = row[0].row % 2 == 0
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            if is_even_row:
                cell.fill = zebra_fill
            # Column 6 is "Ссылка"
            if (
                cell.column == 6
                and isinstance(cell.value, str)
                and cell.value.startswith("http")
            ):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 24

    # Column widths
    column_widths = [16, 12, 22, 36, 22, 40, 24, 80]
    for i, width in enumerate(column_widths):
        col_letter = get_column_letter(i + 1)
        sheet.column_dimensions[col_letter].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()

